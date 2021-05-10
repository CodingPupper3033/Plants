from openpyxl import Workbook, load_workbook

from Garden import Garden
from Plant import Plant

include_maybe = True


class GardenMaker:
    def run(self):
        garden = Garden()
        workbook = load_workbook("Garden.xlsx")
        sheet = workbook.active

        plant_current_name = None
        plant_current_array = []

        for row in range(sheet.min_row + 1, sheet.max_row):
            plant_name = sheet.cell(row, 1).value.lower()
            plant_next_to_value = sheet.cell(row, 3).value.lower()
            plant_next_to_name = sheet.cell(row, 2).value.lower()

            if plant_current_name is None:
                plant_current_name = plant_name

            if plant_current_name != plant_name:
                garden.add_plant(Plant(plant_current_name, plant_current_array))
                plant_current_name = plant_name
                plant_current_array = []

            if plant_next_to_value == "yes" or (plant_next_to_value == "maybe" and include_maybe):
                plant_current_array.append(plant_next_to_name)

        garden.set_plants_next_to_object()
        return garden
